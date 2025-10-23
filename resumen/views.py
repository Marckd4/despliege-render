from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum
from productos.models import Producto
import openpyxl
from openpyxl.utils import get_column_letter

# Datos para la tabla dinámica
def resumen_cod_dun(request):
    resumen = (
        Producto.objects
        .values('cod_dun', 'ubicacion', 'cod_sistema')  # <- agregado cod_sistema
        .annotate(total_cajas=Sum('cajas'))
        .order_by('cod_dun', 'ubicacion')
    )
    return JsonResponse(list(resumen), safe=False)

# Vista del template
def resumen_view(request):
    return render(request, 'resumen/resumen.html')

# Exportación a Excel
def exportar_resumen_excel(request):
    resumen = (
        Producto.objects
        .values('cod_dun', 'ubicacion', 'cod_sistema')  # <- agregado cod_sistema
        .annotate(total_cajas=Sum('cajas'))
        .order_by('cod_dun', 'ubicacion')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen Cod_DUN"

    # Encabezados
    ws.append(["Cod_DUN", "Ubicación", "Cod_Sistema", "Total Cajas"])

    # Datos
    for item in resumen:
        ws.append([item['cod_dun'], item['ubicacion'], item['cod_sistema'], item['total_cajas']])
    
    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Resumen_CodDUN.xlsx"'
    wb.save(response)
    return response
